# -*- coding: utf-8 -*-
# @TIME     :2021/1/11 21:02
# @Author   :Hachi
# @Email    :459327366@qq.com
# @File     :handle_excel.py
# @Software :PyCharm


import openpyxl


class ExcelHandle:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        '读取Excel数据'
        work_book = openpyxl.load_workbook(self.filename)
        sh = work_book[self.sheetname]

        res = list(sh.rows)
        title = [i.value for i in res[0]]
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        '数据写入方法'
        work_book = openpyxl.load_workbook(self.filename)
        sh = work_book[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        work_book.save(self.filename)
